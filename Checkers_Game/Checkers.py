import operator
import numpy as np
import pygame
import sys
import os
import openpyxl

# Clock Object for frame limiting
clock = pygame.time.Clock()
clock.tick(60)
pygame.init()
pygame.font.init()

WIDTH = 800
HEIGHT = 600
screen = pygame.display.set_mode((WIDTH, HEIGHT))

board = np.full((8,8),None, dtype = object)
player1Win = False
player2Win = False
playerEntering = 0
userText = ''

leaderBoardList = []
leaderBoardSortList = ["Wins", "Losses", "Ratio"]
sortCounter = 0
leaderBoardSort = leaderBoardSortList[sortCounter]

# colors
RED = (227, 66, 52); ORANGE = (255, 128, 0); YELLOW = (255, 255, 0); GREEN = (34, 139, 34); BLUE = (0, 0, 255); PURPLE = (255, 0, 255); WHITE = (255, 255, 255); BLACK = (0, 0, 0); BROWN = (150, 75, 0); GRAY = (128,128,128)
BACKGROUND_COLOR = (60, 95, 74)

boardSquare1List = [WHITE, BROWN, RED, RED]
boardSquare2List = [BLACK, BLACK, BROWN, BLACK]
topPieceList = [RED, RED, WHITE, WHITE]
bottomPieceList = [BLACK, BLACK, BLACK, BROWN]
colorCounter = 0

#defult colors
topPieceColor = RED
bottomPieceColor = BLACK
topKingColor = YELLOW
bottomKingColor= GREEN
potenialMovePieceColor = BLUE
boardSquare1Color = boardSquare1List[0]
boardSquare2Color = boardSquare2List[0]

#Assets list
image_path = os.path.join('Assets', 'Black piece.png')
blackPiece = pygame.image.load(image_path).convert_alpha()
image_path = os.path.join('Assets', 'Brown piece.png')
brownPiece = pygame.image.load(image_path).convert_alpha()
image_path = os.path.join('Assets', 'White piece.png')
whitePiece = pygame.image.load(image_path).convert_alpha()
image_path = os.path.join('Assets', 'Red piece.png')
redPiece = pygame.image.load(image_path).convert_alpha()
image_path = os.path.join('Assets', 'Gray piece.png')
grayPiece = pygame.image.load(image_path).convert_alpha()

image_path = os.path.join('Assets', 'Black King.png')
blackKing= pygame.image.load(image_path).convert_alpha()
image_path = os.path.join('Assets', 'Brown King.png')
brownKing= pygame.image.load(image_path).convert_alpha()
image_path = os.path.join('Assets', 'Red King.png')
redKing= pygame.image.load(image_path).convert_alpha()
image_path = os.path.join('Assets', 'White King.png')
whiteKing= pygame.image.load(image_path).convert_alpha()

image_path = os.path.join('Assets', 'Leaderboard Symbol.png')
leaderBoardSymbol= pygame.image.load(image_path).convert_alpha()
image_path = os.path.join('Assets', 'Settings Gear.png')
settingsGearSymbol= pygame.image.load(image_path).convert_alpha()
image_path = os.path.join('Assets', 'Check Mark.png')
checkSymbol= pygame.image.load(image_path).convert_alpha()


lowerImageList = [blackPiece, blackPiece, blackPiece, brownPiece]
upperImageList = [redPiece, redPiece, whitePiece, whitePiece]
lowerKingImageList = [blackKing, blackKing, blackKing, brownKing]
upperKingImageList = [redKing, redKing, whiteKing, whiteKing]


#Importnat mouse postiitons
settingButton = pygame.Rect(725, 25, 50, 50)
colorDown = pygame.Rect(450, 113, 75, 75)
colorUp = pygame.Rect(650, 113, 75, 75)
rematch = pygame.Rect(100, 400, 250, 100)
quitButton = pygame.Rect(450, 400, 250, 100)
toggleStatsButton = pygame.Rect(575, 413, 75, 75)
surrender1 = pygame.Rect(300, 263, 150, 75)
surrender2 = pygame.Rect(550, 263, 150, 75)
leaderBoardButton = pygame.Rect(650, 25, 50, 50)
leaderBoardExitButton = pygame.Rect(25, 25, 50, 50)
sortButton = pygame.Rect(25, 100, 50, 50)

class Piece:
    def __init__(self, location, value):
        self.location = location
        self.x = location[0]
        self.y = location[1]
        self.value = value
        if value == -1:
            self.color = bottomPieceColor
        elif value == 1:
            self.color = topPieceColor
        elif self.value == 99:
            self.color = potenialMovePieceColor
        elif self.value == 5:
            self.color = ORANGE #bottomKingColor
        elif self.value == -5:
            self.color = topKingColor

        self.rects = pygame.Rect(self.location[1] * 50 + 50, self.location[0] * 50 + 50, 50, 50)
        self.is_king = False
    def __str__(self):
        return str(self.value)

    __repr__ = __str__

    def update_loc(self, new_loc):
        self.location = new_loc
        self.x = new_loc[0]
        self.y = new_loc[1]
        self.rects = pygame.Rect(self.location[1] * 50 + 50, self.location[0] * 50 + 50, 50, 50)

    def king(self):
        self.is_king = True
        if self.value == 1:
            self.value = 5
            self.color = bottomKingColor
        elif self.value == -1:
            self.value = -5
            self.color = topKingColor

#used for stats purposes rather than main gameplay
class Player:
    def __init__(self, name):
        self.name = name
        self.moves = 0
        self.captures = 0
        self.kings = 0

#Used for the leaderboard array
class LeadBoardPlayer:
    def __init__(self, name, wins, losses, ratio):
        self.name = name
        self.wins = wins
        self.losses = losses
        self.ratio = ratio

# fill the board with pieces in the right place
# 1 for light and -1 for dark

def instantiateBoard():
    for row in range(8):
        for col in range(8):
            if row == 0 or row == 2:
                if col % 2 == 0:
                    board[row][col] = Piece((row,col),1)
            elif row == 1:
                if col % 2 == 1:
                    board[row][col] = Piece((row,col),1)

            elif row == 5 or row == 7:
                if col % 2 == 1:
                    board[row][col] = Piece((row,col), -1)
            elif row == 6:
                if col % 2 == 0:
                    board[row][col] = Piece((row,col), -1)

def getPossibleMoves(piece, jump_piece):  # 1 is red, -1 is black, 5 is red king, -5 is black king
    # input piece output: list of location tuples
    row = piece.x
    col = piece.y
    possible_moves = []

    # white (1) pieces can only move down i.e. col + 1
    if piece.value == 1:
        if row != 7:
            if col != 0:  # check down left
                if board[row + 1][col - 1] is None:  # no piece in that loc
                    possible_moves.append((row + 1, col - 1))
                elif board[row + 1][col - 1].value == -1 or board[row + 1][col - 1].value == -5:  # down left is a dark piece
                    if col > 1 and row < 6 and board[row + 2][col - 2] is None:
                        possible_moves.append((row + 2, col - 2))

            if col != 7:  # check down right
                if board[row + 1][col + 1] is None:  # no piece in that loc
                    possible_moves.append((row + 1, col + 1))
                elif board[row + 1][col + 1].value == -1 or board[row + 1][col + 1].value == -5:  # down left is a dark piece
                    if row != 6 and col < 6 and board[row + 2][col + 2] is None:
                        possible_moves.append((row + 2, col + 2))
    # dark pieces move up
    if piece.value == -1:
        if row != 0:
            if col != 0:  # check up left
                if board[row - 1][col - 1] is None:  # no piece in that loc
                    possible_moves.append((row - 1, col - 1))
                elif board[row - 1][col - 1].value == 1 or board[row - 1][col - 1].value == 5:  # up left is a light piece
                    if col != 1 and board[row - 2][col - 2] is None:
                        possible_moves.append((row - 2, col - 2))

            if col != 7:  # check up right
                if board[row - 1][col + 1] is None:  # no piece in that loc
                    possible_moves.append((row - 1, col + 1))
                elif board[row - 1][col + 1].value == 1 or board[row - 1][col + 1].value == 5:  # up right is a dark piece
                    if col != 6 and board[row - 2][col + 2] is None:
                        possible_moves.append((row - 2, col + 2))

    if piece.value == 5: #red king
        if row != 0:
            if col != 0:  # check up left
                if board[row - 1][col - 1] is None:  # no piece in that loc
                    possible_moves.append((row - 1, col - 1))
                elif board[row - 1][col - 1].value == -1 or board[row - 1][col - 1].value == -5:  # up left is a light piece
                    if col != 1 and board[row - 2][col - 2] is None:
                        possible_moves.append((row - 2, col - 2))

            if col != 7:  # check up right
                if board[row - 1][col + 1] is None:  # no piece in that loc
                    possible_moves.append((row - 1, col + 1))
                elif board[row - 1][col + 1].value == -1 or board[row - 1][col + 1].value == -5:  # up right is a dark piece
                    if col != 6 and board[row - 2][col + 2] is None:
                        possible_moves.append((row - 2, col + 2))
        if row != 7:
            if col != 0:  # check down left
                if board[row + 1][col - 1] is None:  # no piece in that loc
                    possible_moves.append((row + 1, col - 1))
                elif board[row + 1][col - 1].value == -1 or board[row + 1][col - 1].value == -5:  # down left is a dark piece
                    if col != 1 and board[row + 2][col - 2] is None:
                        possible_moves.append((row + 2, col - 2))

            if col != 7:  # check down right
                if board[row + 1][col + 1] is None:  # no piece in that loc
                    possible_moves.append((row + 1, col + 1))
                elif board[row + 1][col + 1].value == -1 or board[row + 1][col + 1].value == -5:  # down left is a dark piece
                    if col != 6 and board[row + 2][col + 2] is None:
                        possible_moves.append((row + 2, col + 2))

    if piece.value == -5: # black king
        if row != 0:
            if col != 0:  # check up left
                if board[row - 1][col - 1] is None:  # no piece in that loc
                    possible_moves.append((row - 1, col - 1))
                elif board[row - 1][col - 1].value == 1 or board[row - 1][col - 1].value == 5:  # up left is a light piece
                    if col != 1 and board[row - 2][col - 2] is None:
                        possible_moves.append((row - 2, col - 2))

            if col != 7:  # check up right
                if board[row - 1][col + 1] is None:  # no piece in that loc
                    possible_moves.append((row - 1, col + 1))
                elif board[row - 1][col + 1].value == 1 or board[row - 1][col + 1].value == 5:  # up right is a dark piece
                    if col != 6 and board[row - 2][col + 2] is None:
                        possible_moves.append((row - 2, col + 2))
        if row != 7:
            if col != 0:  # check down left
                if board[row + 1][col - 1] is None:  # no piece in that loc
                    possible_moves.append((row + 1, col - 1))
                elif board[row + 1][col - 1].value == 1 or board[row + 1][col - 1].value == 5:  # down left is a dark piece
                    if col != 1 and board[row + 2][col - 2] is None:
                        possible_moves.append((row + 2, col - 2))

            if col != 7:  # check down right
                if board[row + 1][col + 1] is None:  # no piece in that loc
                    possible_moves.append((row + 1, col + 1))
                elif board[row + 1][col + 1].value == 1 or board[row + 1][col + 1].value == 5:  # down left is a dark piece
                    if col != 6 and board[row + 2][col + 2] is None:
                        possible_moves.append((row + 2, col + 2))

    # if the double jump piece is active, remove any potential moves that are not jumps
    if jump_piece is not None:
        temp_list = []
        for loc in possible_moves:
            if (loc[1] - col) % 2 == 0 or (loc[0] - row) % 2 == 0:
                temp_list.append(loc)
        possible_moves = temp_list
    return possible_moves

def move(location, destination):
    # promote to king if necessary
    if board[location[0]][location[1]].value == -1:
        if destination[0] == 0:
            board[location[0]][location[1]].king()
            player2.kings += 1

    if board[location[0]][location[1]].value == 1:
        if destination[0] == 7:
            board[location[0]][location[1]].king()
            player1.kings += 1

    # warning! this method assumes all inputs are valid and in bounds
    # only moves pieces and removes takes
    diff_row = (destination[0] - location[0])
    diff_col = (destination[1] - location[1])

    #if the move is a take
    if diff_row % 2 == 0:
        remove_row = location[0] + (diff_row // 2)
        remove_col = location[1] + (diff_col // 2)
        board[remove_row][remove_col] = None

    # reassigns the location attribute to the moved piece
    board[location].update_loc(destination)
    # moves the element within the board array
    board[destination] = board[location]
    # removes the original piece in the board
    board[location] = None

def drawBoard():
    screen.fill(BACKGROUND_COLOR)
    #create board
    for i in range(0,4):
        for j in range(0,8):
            if j % 2 == 0:
                pygame.draw.rect(screen, boardSquare1Color, (50 + 100 * i, 50 + 50 * j, 50, 50), 0)
                pygame.draw.rect(screen, boardSquare2Color, (100 + 100 * i, 50 + 50 * j, 50, 50), 0)
            else:
                pygame.draw.rect(screen, boardSquare2Color, (50 + 100 * i, 50 + 50 * j, 50, 50), 0)
                pygame.draw.rect(screen, boardSquare1Color, (100 + 100 * i, 50 + 50 * j, 50, 50), 0)

    # Settings button
    pygame.draw.rect(screen, GRAY, (725, 25, 50, 50), 0)
    pygame.draw.rect(screen, BLACK, (725, 25, 50, 50), 3)
    settingsGearScaled = pygame.transform.scale(settingsGearSymbol, (40, 40))
    settingsGearSymbolRect = settingsGearScaled.get_rect()
    settingsGearSymbolRect.center = (750, 50)
    screen.blit(settingsGearScaled, settingsGearSymbolRect)

    # Leaderboard button
    pygame.draw.rect(screen, GRAY, (650, 25, 50, 50), 0)
    pygame.draw.rect(screen, BLACK, (650, 25, 50, 50), 3)
    leaderBoardScaled = pygame.transform.scale(leaderBoardSymbol, (50, 60))
    leaderBoardSymbolRect = leaderBoardScaled.get_rect()
    leaderBoardSymbolRect.center = (675, 40)
    screen.blit(leaderBoardScaled, leaderBoardSymbolRect)

    #Names
    font = pygame.font.SysFont('Arial', 48, bold=True)
    textSurface = font.render(player1.name, True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (650, 100)
    screen.blit(textSurface, textRect)

    textSurface = font.render(player2.name, True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (650, 550)
    screen.blit(textSurface, textRect)

    font = pygame.font.SysFont('Arial', 35)
    #Checkers left
    player1PiecesLeft = 12 - player2.captures
    text = "Pieces Left = " + str(player1PiecesLeft)

    textSurface = font.render(text , True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (650, 150)
    screen.blit(textSurface, textRect)

    player1PiecesLeft = 12 - player1.captures
    text = "Pieces Left = " + str(player1PiecesLeft)

    textSurface = font.render(text, True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (650, 500)
    screen.blit(textSurface, textRect)

    if toggleStats:

        #captures
        text = "Captures: "  + str(player1.captures)
        textSurface = font.render(text, True, BLACK)
        textRect = textSurface.get_rect()
        textRect.center = (650, 200)
        screen.blit(textSurface, textRect)

        text = "Captures: "  + str(player2.captures)
        textSurface = font.render(text, True, BLACK)
        textRect = textSurface.get_rect()
        textRect.center = (650, 450)
        screen.blit(textSurface, textRect)

        #moves
        text = "Moves: " + str(player1.moves)
        textSurface = font.render(text, True, BLACK)
        textRect = textSurface.get_rect()
        textRect.center = (650, 250)
        screen.blit(textSurface, textRect)

        text = "Moves: " + str(player2.moves)
        textSurface = font.render(text, True, BLACK)
        textRect = textSurface.get_rect()
        textRect.center = (650, 400)
        screen.blit(textSurface, textRect)

        #promotions
        text = "Promotions: " + str(player1.kings)
        textSurface = font.render(text, True, BLACK)
        textRect = textSurface.get_rect()
        textRect.center = (650, 300)
        screen.blit(textSurface, textRect)

        text = "Promotions: " + str(player2.kings)
        textSurface = font.render(text, True, BLACK)
        textRect = textSurface.get_rect()
        textRect.center = (650, 350)
        screen.blit(textSurface, textRect)

        pygame.draw.line(screen, BLACK, (525, 325), (775, 325), 5)

    font = pygame.font.SysFont('Arial', 60, bold=True)
    global turn
    if turn == 1:
        text = str(player1.name) + "'s turn"
    else:
        text = str(player2.name) + "'s turn"

    textSurface = font.render(text, True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (250, 500)
    screen.blit(textSurface, textRect)


instantiateBoard()

### Display piece location based on the board object

def returnClickedPiece(pos, board):
    x, y = pos
    if (50 <= x <= 450) & (50 <= y <= 450): #clicked on the board
        i = (x - 50) // 50
        j = (y - 50) // 50
        return board[j][i] # WARNING - can return None

def displayBoardState(board):
    for element in board.flat:
        global topPieceColor
        global bottomPieceColor
        global potenialMovePieceColor
        global topKingColor
        global bottomKingColor

        if element is not None and element.value == 1:
            pygame.draw.circle(screen, topPieceColor , element.rects.center, 15)
            upperScaled = pygame.transform.scale(upperImageList[colorCounter % 4], (50, 50))
            upperRect = upperScaled.get_rect()
            upperRect.center = (element.rects.center[0], element.rects.center[1])
            screen.blit(upperScaled, upperRect)

        if element is not None and element.value == -1:
            lowerScaled = pygame.transform.scale(lowerImageList[colorCounter % 4], (50, 50))
            lowerRect = lowerScaled.get_rect()
            lowerRect.center = (element.rects.center[0], element.rects.center[1])
            screen.blit(lowerScaled, lowerRect)

        if element is not None and element.value == 99:
            grayScaled = pygame.transform.scale(grayPiece, (35, 35))
            grayRect = grayScaled.get_rect()
            grayRect.center = (element.rects.center[0], element.rects.center[1])
            screen.blit(grayScaled, grayRect)

        if element is not None and element.value == 5:
            pygame.draw.circle(screen, topKingColor , element.rects.center, 15)
            upperKingScaled = pygame.transform.scale(upperKingImageList[colorCounter % 4], (50, 50))
            upperKingRect = upperKingScaled.get_rect()
            upperKingRect.center = (element.rects.center[0], element.rects.center[1])
            screen.blit(upperKingScaled, upperKingRect)

        if element is not None and element.value == -5:
            lowerKingScaled = pygame.transform.scale(lowerKingImageList[colorCounter % 4], (50, 50))
            lowerKingRect = lowerKingScaled.get_rect()
            lowerKingRect.center = (element.rects.center[0], element.rects.center[1])
            screen.blit(lowerKingScaled , lowerKingRect)

def loadPotentialMovePieces(loc_list):
    # loads and updates the board with the current list of potential moves

    for i in range(8):
        for j in range(8):
            if possible_moves_board[i][j] is not None:
                if possible_moves_board[i][j].value == 99:
                    possible_moves_board[i][j] = None
    for loc in loc_list:
        x, y = loc
        piece = Piece(loc,99)
        possible_moves_board[x][y]=piece

def checkWin(board):
    red_moves = 0
    black_moves = 0
    global player1Win
    global player2Win
    for element in board.flat:
        if element is not None:
            if getPossibleMoves(element, None):
                if element.value == 1 or element.value == 5:
                    red_moves += 1
                elif element.value == -1 or element.value == -5:
                    black_moves += 1
        if red_moves >= 1 and black_moves >= 1:
            break
    else: # if the board iterates through without both colors having possible moves
        if black_moves == 0 and red_moves > 0:
            player1Win = True
        elif red_moves == 0 and black_moves > 0:
            player2Win = True

def loadLeaderBoard():
    workbook = openpyxl.load_workbook(os.path.join('Assets','Leaderboard.xlsx'))
    sheet = workbook.active
    row = 1
    col = 1
    cell = sheet.cell(row=row, column=col)
    while sheet.cell(row=row, column=1).value is not None:
        name = sheet.cell(row=row, column=1).value

        wins = sheet.cell(row=row, column=2).value

        losses = sheet.cell(row=row, column=3).value

        ratio = sheet.cell(row=row, column=4).value

        tempLeaderBoard = LeadBoardPlayer(name, wins, losses, ratio)
        leaderBoardList.append(tempLeaderBoard)

        row += 1

def updateLeaderBoard():

    player1Unquie = True
    player2Unquie = True

    for element in leaderBoardList:
        if player1.name == element.name:
            player1Unquie = False
            if player1Win:
                element.wins += 1
            else:
                element.losses += 1
            if element.losses == 0:
                element.ratio = element.wins
            else:
                element.ratio = round((element.wins / element.losses), 3)

    for element in leaderBoardList:
        if player2.name == element.name:
            player2Unquie = False
            if player2Win:
                element.wins += 1
            else:
                element.losses += 1

            if element.losses == 0:
                element.ratio = element.wins
            else:
                element.ratio = round((element.wins / element.losses), 3)

    if player1Unquie:
        if player1Win:
            tempPlayer = LeadBoardPlayer(player1.name, 1, 0, 1)
        else:
            tempPlayer = LeadBoardPlayer(player1.name, 0, 1, 0)
        leaderBoardList.append(tempPlayer)

    if player2Unquie:
        if player2Win:
            tempPlayer = LeadBoardPlayer(player2.name, 1, 0, 1)
        else:
            tempPlayer = LeadBoardPlayer(player2.name, 0, 1, 0)
        leaderBoardList.append(tempPlayer)

    workbook = openpyxl.load_workbook(os.path.join('Assets', 'Leaderboard.xlsx'))
    sheet = workbook.active
    row = 1
    col = 1
    for element in leaderBoardList:
        sheet.cell(row=row, column=1).value = element.name

        sheet.cell(row=row, column=2).value = element.wins

        sheet.cell(row=row, column=3).value = element.losses

        sheet.cell(row=row, column=4).value = element.ratio

        row += 1

    # save workbook
    workbook.save(os.path.join('Assets', 'Leaderboard.xlsx'))

def drawSettingsScren():
    pygame.draw.rect(screen, BACKGROUND_COLOR, (0, 0, 800, 600), 0)

    #color chnager section
    pygame.draw.rect(screen, BLACK, (50, 100, 700, 100), 5)
    font = pygame.font.SysFont('Arial', 48)
    textSurface = font.render("Choose Color Pallet", True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (250, 150)
    screen.blit(textSurface, textRect)
    pygame.draw.rect(screen, WHITE, (450, 113, 75, 75), 0)
    pygame.draw.rect(screen, BLACK, (450, 113, 75, 75), 3)
    pygame.draw.rect(screen, WHITE, (650, 113, 75, 75), 0)
    pygame.draw.rect(screen, BLACK, (650, 113, 75, 75), 3)
    pygame.draw.rect(screen, boardSquare1Color, (550, 160, 25, 25), 0)
    pygame.draw.rect(screen, boardSquare2Color, (550, 115, 25, 25), 0)
    pygame.draw.circle(screen, topPieceColor, (605, 170), 15)
    pygame.draw.circle(screen, bottomPieceColor, (605, 125), 15)
    pygame.draw.line(screen, BLACK, (472, 148), (497, 123), 20)
    pygame.draw.line(screen, BLACK, (472, 148), (497, 178), 20)

    pygame.draw.line(screen, BLACK, (699, 148), (674, 123), 20)
    pygame.draw.line(screen, BLACK, (699, 148), (674, 178), 20)

    #surrender button
    pygame.draw.rect(screen, BLACK, (50, 250, 700, 100), 5)
    font = pygame.font.SysFont('Arial', 48)
    textSurface = font.render("Surrender", True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (175, 300)
    screen.blit(textSurface, textRect)
    pygame.draw.rect(screen, WHITE, (300, 263, 150, 75), 0)
    pygame.draw.rect(screen, BLACK, (300, 263, 150, 75), 3)


    pygame.draw.rect(screen, WHITE, (550, 263, 150, 75), 0)
    pygame.draw.rect(screen, BLACK, (550, 263, 150, 75), 3)

    if doubleCheckSurrender1 == False:
        if len(player1.name) > 6:
            font = pygame.font.SysFont('Arial', 30)
        else:
            font = pygame.font.SysFont('Arial', 48)
        text = player1.name
    else:
        font = pygame.font.SysFont('Arial', 30)
        text = "Surrender?"
    textSurface = font.render(text, True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (372, 300)
    screen.blit(textSurface, textRect)

    if doubleCheckSurrender2 == False:
        if len(player2.name) > 6:
            font = pygame.font.SysFont('Arial', 30)
        else:
            font = pygame.font.SysFont('Arial', 48)
        text = player2.name
    else:
        font = pygame.font.SysFont('Arial', 30)
        text = "Surrender?"
    textSurface = font.render(text, True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (622, 300)
    screen.blit(textSurface, textRect)

    #Stats toggle
    pygame.draw.rect(screen, BLACK, (50, 400, 700, 100), 5)
    font = pygame.font.SysFont('Arial', 48)
    textSurface = font.render("Toggle Session Stats", True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (250, 450)
    screen.blit(textSurface, textRect)
    pygame.draw.rect(screen, WHITE, (575, 413, 75, 75), 0)
    pygame.draw.rect(screen, BLACK, (575, 413, 75, 75), 5)

    pygame.draw.rect(screen, GRAY, (725, 25, 50, 50), 0)
    pygame.draw.rect(screen, BLACK, (725, 25, 50, 50), 3)
    pygame.draw.line(screen, RED, (737, 35), (762, 60), 10)
    pygame.draw.line(screen, RED, (737, 60), (762, 35), 10)

    if toggleStats:
        drawCheck()

def drawCheck():
    checkScaled = pygame.transform.scale(checkSymbol, (100, 100))
    checkSymbolRect = checkScaled.get_rect()
    checkSymbolRect.center = (612, 450)
    screen.blit(checkScaled, checkSymbolRect)

def drawStartScreen():
    pygame.draw.rect(screen, BACKGROUND_COLOR, (0, 0, 800, 600), 0)

    #Top Caption
    font = pygame.font.SysFont('Arial', 48*2)
    textSurface = font.render("Welcome to Checkers", True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (400, 100)
    screen.blit(textSurface, textRect)

    #Credits
    font = pygame.font.SysFont('Arial', 48)
    textSurface = font.render("A game by", True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (400, 200)
    screen.blit(textSurface, textRect)
    textSurface = font.render("Lance Fenicle & Jack Meadows", True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (400, 300)
    screen.blit(textSurface, textRect)

    #Continue
    textSurface = font.render("Press Enter to Continue", True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (400, 400)
    screen.blit(textSurface, textRect)

def drawInputScreen():
    pygame.draw.rect(screen, BACKGROUND_COLOR, (0, 0, 800, 600), 0)
    pygame.draw.rect(screen, GRAY, (275, 375, 250, 50), 0)
    pygame.draw.rect(screen, BLACK, (275, 375, 250, 50), 3)

    global userText
    font = pygame.font.SysFont('Arial', 48)
    textSurface = font.render("Please Enter Name:", True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (400, 200)
    screen.blit(textSurface, textRect)

    textSurface = font.render(userText, True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (400, 400)
    screen.blit(textSurface, textRect)

def drawEndScreen():
    pygame.draw.rect(screen, BACKGROUND_COLOR, (0, 0, 800, 600), 0)
    font = pygame.font.SysFont('Arial', 48 * 2)


    text = winner + " Wins!"
    textSurface = font.render(text, True, BLACK)


    textRect = textSurface.get_rect()
    textRect.center = (400, 100)
    screen.blit(textSurface, textRect)

    #Rematch
    pygame.draw.rect(screen, GRAY, (100, 400, 250, 100), 0)
    pygame.draw.rect(screen, BLACK, (100, 400, 250, 100), 3)
    font = pygame.font.SysFont('Arial', 48)
    textSurface = font.render("Rematch", True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (225, 450)
    screen.blit(textSurface, textRect)

    #Quit
    pygame.draw.rect(screen, GRAY, (450, 400, 250, 100), 0)
    pygame.draw.rect(screen, BLACK, (450, 400, 250, 100), 3)
    font = pygame.font.SysFont('Arial', 48)
    textSurface = font.render("End Game", True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (575, 450)
    screen.blit(textSurface, textRect)

def drawLeaderBoardScreen():
    pygame.draw.rect(screen, BACKGROUND_COLOR, (0, 0, 800, 600), 0)

    global leaderBoardSort
    global leaderBoardList
    match leaderBoardSort:
        case "Wins":
            leaderBoardList = sorted(leaderBoardList, key=operator.attrgetter('wins', 'name'))
            leaderBoardList.reverse()
        case "Losses":
            leaderBoardList = sorted(leaderBoardList, key=operator.attrgetter('losses', 'name'))
            leaderBoardList.reverse()
        case "Ratio":
            leaderBoardList = sorted(leaderBoardList, key=operator.attrgetter('ratio', 'name'))
            leaderBoardList.reverse()

    font = pygame.font.SysFont('Arial', 48)
    i = 0

    #heading text
    textSurface = font.render('Name', True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (200, 20)
    screen.blit(textSurface, textRect)

    #win
    textSurface = font.render('Wins', True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (400, 20)
    screen.blit(textSurface, textRect)

    #lose
    textSurface = font.render('Losses', True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (550, 20)
    screen.blit(textSurface, textRect)

    #ratio
    textSurface = font.render('Ratio', True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (700, 20)
    screen.blit(textSurface, textRect)

    while (len(leaderBoardList) > i) and i < 10:

        #name
        text = leaderBoardList[i].name
        textSurface = font.render(text, True, BLACK)
        textRect = textSurface.get_rect()
        textRect.center = (200, 75 + 55 * i)
        screen.blit(textSurface, textRect)

        #win
        text = str(leaderBoardList[i].wins)
        textSurface = font.render(text, True, BLACK)
        textRect = textSurface.get_rect()
        textRect.center = (400, 75 + 55* i)
        screen.blit(textSurface, textRect)

        #lose
        text = str(leaderBoardList[i].losses)
        textSurface = font.render(text, True, BLACK)
        textRect = textSurface.get_rect()
        textRect.center = (550, 75 + 55 * i)
        screen.blit(textSurface, textRect)

        #ratio
        text = str(leaderBoardList[i].ratio)
        textSurface = font.render(text, True, BLACK)
        textRect = textSurface.get_rect()
        textRect.center = (700, 75 + 55 * i)
        screen.blit(textSurface, textRect)

        i += 1
    for i in range(0,10):
        pygame.draw.line(screen, BLACK, (125, 50+ 55 * i), (750, 50+ 55 * i) ,5)

    pygame.draw.rect(screen, GRAY, (25, 25, 50, 50), 0)
    pygame.draw.rect(screen, BLACK, (25, 25, 50, 50), 3)
    pygame.draw.line(screen, RED, (37, 35), (62, 60), 10)
    pygame.draw.line(screen, RED, (37, 60), (62, 35), 10)

    pygame.draw.rect(screen, WHITE, (25, 100, 50, 50), 0)
    pygame.draw.rect(screen, BLACK, (25, 100, 50, 50), 3)
    pygame.draw.line(screen, BLACK, (34, 112), (49, 137), 10)
    pygame.draw.line(screen, BLACK, (49, 137), (64, 112), 10)

    if leaderBoardSort == "Losses":
        font = pygame.font.SysFont('Arial', 40)
    else:
        font = pygame.font.SysFont('Arial', 48)

    text = leaderBoardSort
    textSurface = font.render(text, True, BLACK)
    textRect = textSurface.get_rect()
    textRect.center = (50, 200)
    screen.blit(textSurface, textRect)


# create list of possible moves
potential_move_pieces = []
#create possible moves board (to be displayed later)
possible_moves_board = np.full((8,8),None, dtype = object)

selected_piece = None # blue potential piece
turn = 1 # default starts with red
double_jump = None

#Screen state boolean
displaySettings = False
displayInput = False
displayEnd = False
displayLeaderboard = False
displayStart = True
displayBoard = False
toggleStats = False
doubleCheckSurrender1 = False
doubleCheckSurrender2 = False
update = False
winner = ''
loadLeaderBoard()

running = True
while running:

    for event in pygame.event.get(): # quit out of the game
        if event.type == pygame.QUIT:
            pygame.quit()
            sys.exit()
        if event.type == pygame.KEYDOWN and displayInput:

            #Input screen logic
            if playerEntering == 0:
                if event.key == pygame.K_BACKSPACE:
                    userText = userText[:-1]  # Remove last character
                elif event.key == pygame.K_RETURN and userText != '':
                    # Do something with the entered text (e.g., process it)
                    player1 = Player(userText)
                    playerEntering = 1
                    userText = ''
                else:
                    if len(userText) < 10:
                        userText += event.unicode  # Add the pressed character

            elif playerEntering == 1:
                if event.key == pygame.K_BACKSPACE:
                    userText = userText[:-1]  # Remove last character
                elif event.key == pygame.K_RETURN and userText != '':
                    # Do something with the entered text (e.g., process it)
                    player2 = Player(userText)
                    displayInput = False
                    displayBoard = True
                    playerEntering = 1
                    userText = ''
                else:
                    if len(userText) < 10:
                        userText += event.unicode  # Add the pressed character

            # Start screen logic
        if event.type == pygame.KEYDOWN and displayStart:
            if event.key == pygame.K_RETURN and displayStart == True:
                displayStart = False
                displayInput = True


        if event.type == pygame.MOUSEBUTTONDOWN:

            mouse_pos = pygame.mouse.get_pos()

            #Settings screen buttons
            if settingButton.collidepoint(mouse_pos):
                if displaySettings:
                    displaySettings = False
                else:
                    displaySettings = True

            if leaderBoardButton.collidepoint(mouse_pos):
                if displayBoard:
                    displayLeaderboard = True
            if displayLeaderboard and leaderBoardExitButton.collidepoint(mouse_pos):
                displayLeaderboard = False

            if displayLeaderboard and sortButton.collidepoint(mouse_pos):
                sortCounter += 1
                leaderBoardSort = leaderBoardSortList[sortCounter % 3]

            #Color changer logic
            if displaySettings and colorUp.collidepoint(mouse_pos):
                colorCounter += 1
                boardSquare1Color = boardSquare1List[colorCounter % 4]
                boardSquare2Color = boardSquare2List[colorCounter % 4]
                topPieceColor = topPieceList[colorCounter % 4]
                bottomPieceColor = bottomPieceList[colorCounter % 4]

            if displaySettings and colorDown.collidepoint(mouse_pos):
                colorCounter -= 1
                boardSquare1Color = boardSquare1List[colorCounter % 4]
                boardSquare2Color = boardSquare2List[colorCounter % 4]
                topPieceColor = topPieceList[colorCounter % 4]
                bottomPieceColor = bottomPieceList[colorCounter % 4]

            #Toggle stats logic
            if displaySettings and toggleStatsButton.collidepoint(mouse_pos):
                if toggleStats == True:
                    toggleStats = False
                else:
                    toggleStats = True

            # Surrender logic
            if displaySettings and surrender1.collidepoint(mouse_pos):
                if doubleCheckSurrender1 == False:
                    doubleCheckSurrender1 = True
                else:
                    displaySettings = False
                    displayBoard = False
                    player2Win = True
            if displaySettings and surrender2.collidepoint(mouse_pos):
                if doubleCheckSurrender2 == False:
                    doubleCheckSurrender2 = True
                else:
                    displaySettings = False
                    displayBoard = False
                    player1Win = True

            #End screen logic
            if displayEnd and rematch.collidepoint(mouse_pos):
                displayEnd = False
                displayStart = True
                player1Win = False
                player2Win = False
                clicked_piece = None
                playerEntering = 0
                instantiateBoard()


            if displayEnd and quitButton.collidepoint(mouse_pos):
                pygame.quit()
                sys.exit()

            clicked_piece = returnClickedPiece(mouse_pos,
                                               board)  # returns the selected piece based off of the given location

            if double_jump is None: # if no double jump continue normally
                if clicked_piece is not None:

                    if clicked_piece.value == turn or clicked_piece.value == (turn * 5): #checks turn
                        loadPotentialMovePieces(getPossibleMoves(clicked_piece, double_jump))

                        if clicked_piece.value != 99: # if the selected piece is not a potential move piece
                            selected_piece = clicked_piece # set the selected piece
            else: # if double jump is active
                # set the selected piece to only the jump piece
                selected_piece = double_jump
                loadPotentialMovePieces(getPossibleMoves(double_jump, double_jump))

                # provide a way to get out of the double jump
                possible_moves = getPossibleMoves(double_jump, double_jump)
                if mouse_pos[0] > 450 or mouse_pos[1] > 450 or len(possible_moves) == 0: # if there are no possible jumps or the player clicks outside the board
                    turn *= -1 # switch the turn

                    # clear
                    possible_moves_board[:] = None
                    selected_piece = None
                    possible_clicked_piece = None
                    double_jump = None

            # if the selected piece was not assigned to the selected piece, assign it to the potential piece
            possible_clicked_piece = returnClickedPiece(mouse_pos, possible_moves_board)

            if possible_clicked_piece is not None and selected_piece is not None: # if both selected and potential are not None

                if (possible_clicked_piece.location[0] - selected_piece.location[0]) % 2 == 0: # determine if the move was a jump
                    double_jump = selected_piece # select the double jump

                    if turn == 1:
                        player1.captures += 1
                    else:
                        player2.captures += 1
                else:
                    double_jump = None # clear the double jump

                move(selected_piece.location, possible_clicked_piece.location) # do the move

                if double_jump is None: # update the turn if not double jump

                    if turn == 1:
                        player1.moves += 1
                    else:
                        player2.moves += 1

                    turn *= -1

                # clear possible moves
                possible_moves_board[:] = None
                # deselect piece
                selected_piece = None

    #display states handled down here
    if displayBoard:
        drawBoard()  # display checkerboard

        displayBoardState(board)  # display checker pieces based on their location in the board array
        displayBoardState(possible_moves_board)  # displays any possible pieces

    if displayStart:
        drawStartScreen()

    if displayInput:
        drawInputScreen()

    if displaySettings:
        drawSettingsScren()

    if displayLeaderboard:
        drawLeaderBoardScreen()

    if player1Win == True or player2Win == True:
        if player1Win:
            winner = player1.name
        else:
            winner = player2.name
        displayEnd = True
        updateLeaderBoard()
        player1Win = False
        player2Win = False
        doubleCheckSurrender1 = False
        doubleCheckSurrender2 = False

    if displayEnd:
        drawEndScreen()

    pygame.display.update()
    checkWin(board)
